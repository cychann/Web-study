import pyautogui
import sys
from openpyxl import load_workbook

wb = load_workbook('NameList.xlsx') 
ws  = wb.active

# 페이지 활성화
w = pyautogui.getWindowsWithTitle('아이템 보기')[0]
if w.isActive == False:
    w.activate()
    pyautogui.sleep(1)

i = 1
while True:
    i += 1

    # 값 확인
    if ws['A{}'.format(i)].value == '복수파일명':
        break
    print(ws['A{}'.format(i)].value)

    # 아이템 추가 버튼 클릭
    pyautogui.sleep(1)
    add_item = pyautogui.locateOnScreen('add_item.png')
    pyautogui.click(add_item)

    # 엑셀 활성화
    w = pyautogui.getWindowsWithTitle('NameList')[0]
    if w.isActive == False:
        w.activate()
        pyautogui.sleep(1)

    # 방향키 DOWN
    pyautogui.press('down')

    # 엑셀 복사
    pyautogui.hotkey('ctrl', 'c')
    pyautogui.sleep(1)

    # 페이지 활성화
    w = pyautogui.getWindowsWithTitle('아이템 추가')[0]
    if w.isActive == False:
        w.activate()
        pyautogui.sleep(1)

    # 제목(필수입력 항목) 클릭
    title = pyautogui.locateOnScreen('title.png')
    pyautogui.click(title)

    # 엑셀 붙여넣기
    pyautogui.hotkey('ctrl', 'v')

    # 파일 이미지 클릭
    file = pyautogui.locateOnScreen('file.png')
    pyautogui.click(file)

    # 파일 선택 이미지 클릭
    select_file = pyautogui.locateOnScreen('select_file.png')
    pyautogui.click(select_file)
    pyautogui.sleep(2)

    # 붙여넣기 엔터
    pyautogui.keyDown('ctrl')
    pyautogui.press('v')
    pyautogui.keyUp('ctrl')
    pyautogui.press('enter')

    # 컬렉션클릭
    select_below = pyautogui.locateOnScreen('select_below.png')
    pyautogui.click(select_below)
    pyautogui.sleep(2)
    collection = pyautogui.locateOnScreen('collection.png')
    pyautogui.click(collection)

    # 공개 체크 박스 클릭
    checkbox = pyautogui.locateOnScreen('checkbox.png')
    pyautogui.click(checkbox)

    # 마지막 아이템 추가 버튼 클릭
    add_item_final = pyautogui.locateOnScreen('add_item_final.png')
    pyautogui.click(add_item_final)
    pyautogui.sleep(5)

i -= 1

# 엑셀 활성화
w = pyautogui.getWindowsWithTitle('NameList')[0]
if w.isActive == False:
    w.activate()
    pyautogui.sleep(2)

# 방향키 DOWN
pyautogui.press('down')
pyautogui.sleep(5)
i += 1

# 페이지 활성화
w = pyautogui.getWindowsWithTitle('아이템 보기')[0]
if w.isActive == False:
    w.activate()
    pyautogui.sleep(1)

while True:
    # <복수 파일>

    # 아이템 추가 버튼 클릭
    pyautogui.sleep(1)
    add_item = pyautogui.locateOnScreen('add_item.png')
    pyautogui.click(add_item)

    # 엑셀 활성화
    w = pyautogui.getWindowsWithTitle('NameList')[0]
    if w.isActive == False:
        w.activate()
        pyautogui.sleep(2)

    # 방향키 DOWN
    pyautogui.press('down')

    i += 1

    # 값 확인
    if ws['A{}'.format(i)].value == 'Finish':
        break

    # 엑셀 복사
    pyautogui.hotkey('ctrl', 'c')
    pyautogui.sleep(1)

    # 페이지 활성화
    w = pyautogui.getWindowsWithTitle('아이템 추가')[0]
    if w.isActive == False:
        w.activate()
        pyautogui.sleep(1)

    # 제목(필수입력 항목) 클릭
    title = pyautogui.locateOnScreen('title.png')
    pyautogui.click(title)

    # 엑셀 붙여넣기
    pyautogui.hotkey('ctrl', 'v')

    # 파일 이미지 클릭
    file = pyautogui.locateOnScreen('file.png')
    pyautogui.click(file)
    pyautogui.sleep(1)

    while True:
        # 엑셀 활성화
        w = pyautogui.getWindowsWithTitle('NameList')[0]
        if w.isActive == False:
            w.activate()
            pyautogui.sleep(2)

        # 방향키 DOWN
        pyautogui.press('down')
        pyautogui.sleep(1)

        i += 1

        print(ws['A{}'.format(i)].value)
        # 값 확인
        if ws['A{}'.format(i)].value == None:
            break
        

        # 엑셀 복사
        pyautogui.hotkey('ctrl', 'c')
        pyautogui.sleep(1)

        # 페이지 활성화
        w = pyautogui.getWindowsWithTitle('아이템 추가')[0]
        if w.isActive == False:
            w.activate()
            pyautogui.sleep(1)

        # 파일 선택 이미지 클릭
        no_selected_file = pyautogui.locateOnScreen('no_selected_file.png')
        pyautogui.click(no_selected_file)
        pyautogui.sleep(2)

        # 붙여넣기 엔터
        pyautogui.hotkey('ctrl', 'v')
        pyautogui.press('enter')
        pyautogui.sleep(1)

        # 다른 파일 선택 이미지 클릭
        other_file_select = pyautogui.locateOnScreen('other_file_select.png')
        pyautogui.click(other_file_select)
        pyautogui.sleep(1)


    # 컬렉션클릭
    select_below = pyautogui.locateOnScreen('select_below.png')
    pyautogui.click(select_below)

    pyautogui.sleep(2)

    collection = pyautogui.locateOnScreen('collection.png')
    pyautogui.click(collection)

    # 공개 체크 박스 클릭
    checkbox = pyautogui.locateOnScreen('checkbox.png')
    pyautogui.click(checkbox)

    # 마지막 아이템 추가 버튼 클릭
    add_item_final = pyautogui.locateOnScreen('add_item_final.png')
    pyautogui.click(add_item_final)
    pyautogui.sleep(5)

sys.exit()

